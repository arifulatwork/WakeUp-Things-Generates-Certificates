# app.py
from flask import Flask, request, send_file, render_template_string, jsonify
import pandas as pd
import openpyxl
from docxtpl import DocxTemplate
from werkzeug.utils import secure_filename
import os
import zipfile
import io
import re
import shutil
from datetime import datetime
import tempfile
import subprocess

app = Flask(__name__)

# Cap upload size to avoid huge/abusive uploads taking down the server.
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024  # 25 MB

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
OUTPUT_FOLDER = os.path.join(os.path.dirname(__file__), "generated")
TEMPLATES_FOLDER = os.path.join(os.path.dirname(__file__), "templates")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(TEMPLATES_FOLDER, exist_ok=True)

# --- Certificate type definitions -----------------------------------------
VET_TEMPLATE_FILE = "certificate_template.docx"
JOB_SHADOWING_FILE = "Job_Shadowing_Certificates.docx"
PORTUGUESE_FILE = "Portuguese_certificate_Template.docx"
TEACHERS_FILE = "teachers_CERTIFICATES .docx"

TEMPLATE_SUFFIX = {
    VET_TEMPLATE_FILE: "VET",
    JOB_SHADOWING_FILE: "Job_Shadowing",
    PORTUGUESE_FILE: "Portuguese_Language",
    TEACHERS_FILE: "Teachers_Certificate",
}

CERT_TYPE_OPTIONS = [
    {"file": VET_TEMPLATE_FILE, "label": "VET Certificate"},
    {"file": JOB_SHADOWING_FILE, "label": "Job Shadowing"},
    {"file": PORTUGUESE_FILE, "label": "Portuguese Language"},
    {"file": TEACHERS_FILE, "label": "Teachers"},
]


def save_upload(file_storage):
    """Save an uploaded file under a sanitized name inside UPLOAD_FOLDER and
    return its path. Using secure_filename() prevents path traversal via a
    crafted filename (e.g. '../../app.py')."""
    filename = secure_filename(file_storage.filename) or "upload.xlsx"
    path = os.path.join(UPLOAD_FOLDER, filename)
    file_storage.save(path)
    return path


# --- DOCX -> PDF conversion (LibreOffice, cross-platform) ------------------
#
# Replaces the old Word COM automation (Windows + MS Word only) with
# LibreOffice headless conversion, which works the same way on Windows,
# Linux, and Mac, needs no Office license, and can convert every generated
# certificate in a single process call instead of one Word launch per file.

def find_libreoffice():
    """Locate the LibreOffice 'soffice' binary across platforms."""
    candidate = shutil.which("soffice") or shutil.which("libreoffice")
    if candidate:
        return candidate

    # The Windows installer doesn't always add soffice.exe to PATH.
    win_candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for path in win_candidates:
        if os.path.exists(path):
            return path

    # Default Mac app bundle location.
    mac_candidate = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    if os.path.exists(mac_candidate):
        return mac_candidate

    return None


LIBREOFFICE_BIN = find_libreoffice()


def convert_docx_to_pdf_batch(docx_paths, output_dir, timeout=300):
    """Convert a batch of .docx files to PDF in a single LibreOffice headless
    call. Cross-platform (Windows/Linux/Mac) as long as LibreOffice is
    installed. Returns {docx_path: pdf_path} for every file that converted
    successfully.
    """
    if not docx_paths:
        return {}

    if not LIBREOFFICE_BIN:
        raise RuntimeError(
            "LibreOffice not found. Install it so the 'soffice' binary is "
            "available: 'sudo apt install libreoffice' (Linux), "
            "'brew install --cask libreoffice' (Mac), or download the "
            "installer from libreoffice.org (Windows)."
        )

    # Give this conversion its own LibreOffice user profile so concurrent
    # requests on the server don't collide on a shared profile lock.
    profile_dir = tempfile.mkdtemp(prefix="lo_profile_")
    try:
        cmd = [
            LIBREOFFICE_BIN,
            "--headless",
            "--norestore",
            f"-env:UserInstallation=file://{profile_dir}",
            "--convert-to", "pdf",
            "--outdir", output_dir,
        ] + docx_paths
        subprocess.run(cmd, capture_output=True, timeout=timeout, check=True)
    finally:
        shutil.rmtree(profile_dir, ignore_errors=True)

    results = {}
    for docx_path in docx_paths:
        base = os.path.splitext(os.path.basename(docx_path))[0]
        pdf_path = os.path.join(output_dir, base + ".pdf")
        if os.path.exists(pdf_path):
            results[docx_path] = pdf_path
    return results


def merge_pdfs(pdf_paths, output_path):
    """Merge multiple PDF files into one using PyPDF2."""
    try:
        from PyPDF2 import PdfMerger
        merger = PdfMerger()
        for pdf in pdf_paths:
            if os.path.exists(pdf):
                merger.append(pdf)
        merger.write(output_path)
        merger.close()
        return True
    except ImportError:
        # If PyPDF2 is not installed, fallback to just returning the first PDF
        if pdf_paths and os.path.exists(pdf_paths[0]):
            shutil.copy(pdf_paths[0], output_path)
            return True
    except Exception:
        pass
    return False


HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Wake Up Projects · Certificate Generator</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600;9..144,700&family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
  :root {
    --navy: #16213E;
    --navy-deep: #0E1530;
    --navy-light: #2A3A5C;
    --amber: #FF8C42;
    --gold: #FFC857;
    --gold-light: #FFE4B5;
    --cream: #FBF6EF;
    --ink: #20232B;
    --muted: #6E7480;
    --muted-light: #A0A6B2;
    --line: #EAE3D8;
    --success: #2E7D32;
    --danger: #C62828;
    --shadow: 0 20px 60px rgba(22,33,62,0.15);
    --radius: 20px;
  }
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif; 
         background: linear-gradient(135deg, #F5F0EB 0%, var(--cream) 100%);
         min-height: 100vh; display: flex; align-items: center; justify-content: center; 
         padding: 24px; color: var(--ink); }

  .card { background: #fff; border-radius: var(--radius); box-shadow: var(--shadow);
          width: 100%; max-width: 720px; overflow: hidden;
          animation: rise 0.6s cubic-bezier(0.16, 1, 0.3, 1); }
  @keyframes rise { from { opacity: 0; transform: translateY(30px) scale(0.98); } 
                    to { opacity: 1; transform: translateY(0) scale(1); } }

  .banner { position: relative; padding: 40px 40px 32px; text-align: center; overflow: hidden;
            background: linear-gradient(160deg, var(--navy-deep) 0%, var(--navy) 60%, #1E3A5F 100%); }
  .banner::after { content: ""; position: absolute; inset: 0; 
                    background: radial-gradient(ellipse at 80% 0%, rgba(255,140,66,0.15) 0%, transparent 70%);
                    pointer-events: none; }
  .banner .logo-wrap {  position: relative;
    z-index: 1;
    width: 120px;
    height: 120px;
    margin: 0 auto 16px;
    border-radius: 50%;
    background: #fff;
    display: flex;
    align-items: center;
    justify-content: center;
    box-shadow: 0 8px 24px rgba(0,0,0,0.3);
    overflow: hidden; }
  .banner .logo-wrap img { width: 90%;
    height: 90%;
    object-fit: contain;}
  .banner .logo-wrap .fallback { font-family: 'Fraunces', serif; font-weight: 700; color: var(--navy); font-size: 24px; }
  .banner h1 { position: relative; z-index: 1; font-family: 'Fraunces', serif; font-weight: 700;
               font-size: 28px; letter-spacing: 0.5px; color: #fff; margin-bottom: 4px; }
  .banner p.sub { position: relative; z-index: 1; font-size: 13px; letter-spacing: 2px;
                  text-transform: uppercase; color: rgba(255,255,255,0.7); font-weight: 500; }

  .body-pad { padding: 32px 40px 28px; }

  .step-indicator { display: flex; justify-content: space-between; margin: 0 0 20px 0;
                    padding: 0 4px; position: relative; }
  .step-indicator::before { content: ""; position: absolute; top: 50%; left: 20px; right: 20px;
                            height: 2px; background: var(--line); transform: translateY(-50%); }
  .step-indicator .step { display: flex; flex-direction: column; align-items: center; gap: 4px;
                          position: relative; z-index: 1; background: white; padding: 0 6px; }
  .step-indicator .step .num { width: 28px; height: 28px; border-radius: 50%; background: var(--line);
                               color: var(--muted); font-size: 12px; font-weight: 700;
                               display: flex; align-items: center; justify-content: center;
                               transition: all 0.3s; }
  .step-indicator .step .label { font-size: 10px; color: var(--muted-light); font-weight: 500;
                                 letter-spacing: 0.3px; text-transform: uppercase; }
  .step-indicator .step.active .num { background: linear-gradient(135deg, var(--amber), var(--gold));
                                       color: white; box-shadow: 0 4px 12px rgba(255,140,66,0.3); }
  .step-indicator .step.active .label { color: var(--navy); font-weight: 600; }
  .step-indicator .step.done .num { background: var(--success); color: white; }

  .drop-zone { border: 2.5px dashed #D5C8B8; border-radius: 16px; padding: 40px 20px;
               text-align: center; cursor: pointer; transition: all 0.3s; background: #FDFBF9;
               position: relative; }
  .drop-zone.drag-over { border-color: var(--amber); background: #FFF8F0; 
                          transform: scale(1.01); box-shadow: 0 8px 24px rgba(255,140,66,0.1); }
  .drop-zone input[type=file] { position: absolute; inset: 0; opacity: 0; cursor: pointer; }
  .drop-zone .icon { font-size: 48px; margin-bottom: 8px; display: block; }
  .drop-zone p { font-size: 15px; color: var(--muted); font-weight: 500; }
  .drop-zone .hint { font-size: 12.5px; color: var(--muted-light); margin-top: 4px; }
  #file-name { margin-top: 14px; font-size: 14px; color: var(--navy); font-weight: 600;
               min-height: 24px; text-align: center; padding: 8px 16px; background: #F0F0F0;
               border-radius: 8px; display: inline-block; }

  .checkbox-group { margin: 16px 0; padding: 16px 20px; background: #FDFBF9; 
                     border-radius: 12px; border: 1.5px solid #EDE6DE; }
  .checkbox-group .group-title { display: block; margin-bottom: 10px; font-weight: 600;
                                  font-size: 13px; color: var(--navy); letter-spacing: 0.3px; }
  .template-options { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; }
  .template-options label { display: flex; align-items: center; gap: 8px; font-size: 13px; 
                            padding: 10px 14px; background: white; border-radius: 10px; 
                            border: 1.5px solid #EDE6DE; cursor: pointer; transition: all 0.2s;
                            font-weight: 500; color: var(--muted); }
  .template-options label:hover { border-color: var(--amber); background: #FFF8F0; }
  .template-options label.checked { border-color: var(--amber); background: #FFF8F0; color: var(--navy); }
  .template-options input[type=checkbox] { accent-color: var(--amber); width: 16px; height: 16px; 
                                           flex-shrink: 0; cursor: pointer; }

  /* Portuguese hours input styling */
  #portuguese-hours-box {
    margin-top: 10px;
    padding: 12px 14px;
    background: #FFF8F0;
    border-radius: 10px;
    border: 1.5px solid #EDE6DE;
    display: none;
  }
  #portuguese-hours-box label {
    font-weight: 600;
    font-size: 13px;
    color: var(--navy);
    display: block;
    margin-bottom: 4px;
  }
  #portuguese-hours-box input {
    width: 100%;
    padding: 10px 14px;
    border-radius: 8px;
    border: 1.5px solid #EDE6DE;
    font-size: 14px;
    font-family: 'Inter', sans-serif;
    background: white;
    transition: all 0.2s;
  }
  #portuguese-hours-box input:focus {
    outline: none;
    border-color: var(--amber);
    box-shadow: 0 0 0 4px rgba(255,140,66,0.1);
  }

  /* Output format selection */
  .format-group {
    margin: 16px 0;
    padding: 16px 20px;
    background: #FDFBF9;
    border-radius: 12px;
    border: 1.5px solid #EDE6DE;
  }
  .format-group .group-title {
    display: block;
    margin-bottom: 10px;
    font-weight: 600;
    font-size: 13px;
    color: var(--navy);
    letter-spacing: 0.3px;
  }
  .format-options {
    display: flex;
    gap: 16px;
    flex-wrap: wrap;
  }
  .format-options label {
    display: flex;
    align-items: center;
    gap: 8px;
    font-size: 13px;
    padding: 8px 14px;
    background: white;
    border-radius: 10px;
    border: 1.5px solid #EDE6DE;
    cursor: pointer;
    transition: all 0.2s;
    font-weight: 500;
    color: var(--muted);
  }
  .format-options label:hover { border-color: var(--amber); background: #FFF8F0; }
  .format-options label.checked { border-color: var(--amber); background: #FFF8F0; color: var(--navy); }
  .format-options input[type=radio] { accent-color: var(--amber); width: 16px; height: 16px; 
                                      flex-shrink: 0; cursor: pointer; }

  .info-box { background: linear-gradient(135deg, #FFF8F0, #FFF3E8); border-left: 4px solid var(--amber);
              border-radius: 10px; padding: 14px 18px; margin: 16px 0; font-size: 13px; 
              color: #5B4D3A; line-height: 1.7; }
  .info-box strong { color: var(--navy); }
  .info-box code { background: rgba(255,140,66,0.1); padding: 2px 8px; border-radius: 4px;
                   font-size: 12px; color: var(--navy); font-family: 'Inter', monospace; }

  .sheet-select-wrap { margin: 16px 0; }
  .sheet-select-wrap label { display: block; font-weight: 600; font-size: 13px; color: var(--navy);
                             margin-bottom: 6px; letter-spacing: 0.3px; }
  .sheet-select-wrap select { width: 100%; padding: 12px 16px; border-radius: 10px;
                              border: 1.5px solid #EDE6DE; font-size: 14px; font-family: 'Inter', sans-serif;
                              background: #FDFBF9; color: var(--ink); transition: all 0.2s;
                              appearance: none; background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='8' viewBox='0 0 12 8'%3E%3Cpath d='M1 1l5 5 5-5' stroke='%236E7480' stroke-width='2' fill='none' stroke-linecap='round'/%3E%3C/svg%3E");
                              background-repeat: no-repeat; background-position: right 16px center; cursor: pointer; }
  .sheet-select-wrap select:focus { outline: none; border-color: var(--amber); 
                                    box-shadow: 0 0 0 4px rgba(255,140,66,0.1); }
  .sheet-select-wrap select:disabled { opacity: 0.5; cursor: not-allowed; }

  button[type=submit] { width: 100%; padding: 16px; background: linear-gradient(135deg, var(--amber), var(--gold));
                        color: var(--navy-deep); border: none; border-radius: 12px; font-size: 16px;
                        font-weight: 700; font-family: 'Inter', sans-serif; cursor: pointer;
                        transition: all 0.25s; margin-top: 8px; letter-spacing: 0.5px;
                        box-shadow: 0 4px 16px rgba(255,140,66,0.3); }
  button[type=submit]:hover:not(:disabled) { transform: translateY(-2px); 
                                              box-shadow: 0 8px 24px rgba(255,140,66,0.4); }
  button[type=submit]:active:not(:disabled) { transform: translateY(0px); }
  button[type=submit]:disabled { background: #E8E0D8; color: #B0A89E; cursor: not-allowed;
                                 box-shadow: none; }

  #progress { display: none; margin-top: 20px; }
  .bar-wrap { background: #F0EDE8; border-radius: 99px; height: 8px; overflow: hidden; }
  .bar { height: 100%; background: linear-gradient(90deg, var(--amber), var(--gold));
         border-radius: 99px; animation: indeterminate 1.8s infinite ease-in-out; }
  @keyframes indeterminate {
    0%   { transform: translateX(-100%); width: 50%; }
    100% { transform: translateX(200%); width: 50%; }
  }
  #status { font-size: 14px; color: var(--muted); margin-top: 10px; text-align: center; font-weight: 500; }

  #result { display: none; margin-top: 24px; text-align: center; }
  #result .success-icon { font-size: 48px; display: block; margin-bottom: 8px; }
  #result .success { color: var(--success); font-size: 18px; font-weight: 700; margin-bottom: 12px; }
  #result .dl-btn { display: inline-flex; align-items: center; gap: 10px; padding: 14px 36px; 
                    background: var(--navy); color: #fff; border-radius: 12px; text-decoration: none; 
                    font-weight: 700; font-size: 15px; transition: all 0.25s; box-shadow: 0 4px 16px rgba(22,33,62,0.25); }
  #result .dl-btn:hover { background: var(--navy-deep); transform: translateY(-2px);
                          box-shadow: 0 8px 24px rgba(22,33,62,0.35); }
  #result .dl-btn .icon { font-size: 20px; }

  #error-box { display: none; margin-top: 16px; padding: 14px 18px; background: #FFF0EE;
               border-left: 4px solid var(--danger); border-radius: 10px; font-size: 14px; 
               color: var(--danger); font-weight: 500; }

  .legend { font-size: 11.5px; color: var(--muted-light); margin-top: 8px; font-style: italic; }

  footer { padding: 18px 40px 22px; text-align: center; border-top: 1.5px solid var(--line); }
  footer p { font-size: 12px; color: var(--muted-light); letter-spacing: 0.3px; }
  footer p strong { color: var(--navy); font-weight: 600; }

  @media (max-width: 600px) {
    .banner { padding: 28px 20px 24px; }
    .body-pad { padding: 20px 18px 24px; }
    .template-options { grid-template-columns: 1fr; }
    .format-options { flex-direction: column; }
    .step-indicator .step .label { font-size: 8px; }
    .step-indicator .step .num { width: 22px; height: 22px; font-size: 10px; }
    footer { padding: 14px 18px 18px; }
  }
</style>
</head>
<body>
<div class="card">
  <div class="banner">
    <div class="logo-wrap">
      <img src="https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcTtaoCdUwAIsKpuSX0CQ5JJpijG6nb44TDKrA&s"
           alt="Wake Up Projects logo"
           onerror="this.style.display='none'; this.nextElementSibling.style.display='flex';">
      <div class="fallback" style="display:none; width:100%; height:100%; align-items:center; justify-content:center;">WP</div>
    </div>
    <h1>Wake Up Projects</h1>
    <p class="sub">Certificate Generator</p>
  </div>

  <div class="body-pad">

  <div class="step-indicator">
    <div class="step active"><span class="num">1</span><span class="label">Configure</span></div>
    <div class="step"><span class="num">2</span><span class="label">Upload File</span></div>
    <div class="step"><span class="num">3</span><span class="label">Download</span></div>
  </div>

  <form id="upload-form" enctype="multipart/form-data">

    <div class="sheet-select-wrap" id="sheet-picker" style="display:none;">
      <label for="sheet-select">📌 Select Group / Sheet</label>
      <select name="sheet_name" id="sheet-select" disabled>
        <option value="">Loading sheets…</option>
      </select>
    </div>

    <div class="checkbox-group">
      <span class="group-title">📜 Select Certificate Types</span>
      <div class="template-options">
        <label>
          <input type="checkbox" name="cert_templates" value="__VET_FILE__" checked>
          VET Certificate
        </label>
        <label>
          <input type="checkbox" name="cert_templates" value="__JOB_SHADOWING_FILE__" checked>
          Job Shadowing
        </label>
        <label>
          <input type="checkbox" name="cert_templates" value="__PORTUGUESE_FILE__" id="portuguese-cert" checked>
          Portuguese Language
        </label>
        <label>
          <input type="checkbox" name="cert_templates" value="__TEACHERS_FILE__" checked>
          Teachers
        </label>
      </div>

      <!-- Portuguese hours input - shows only when Portuguese checkbox is checked -->
      <div id="portuguese-hours-box">
        <label for="portuguese_hours">🇵🇹 Course Hours</label>
        <input type="number"
               name="portuguese_hours"
               id="portuguese_hours"
               min="1"
               max="999"
               value="6"
               placeholder="Enter course hours (default: 6)">
        <div class="legend" style="margin-top:4px;">Hours to display on Portuguese language certificates</div>
      </div>

      <div class="legend">✓ Teachers certificates only generate for rows where Sector = "School" in the selected group</div>
    </div>

    <!-- Output Format Selection -->
    <div class="format-group">
      <span class="group-title">📄 Output Format</span>
      <div class="format-options">
        <label class="checked">
          <input type="radio" name="output_format" value="pdf" checked>
          PDF (Merged)
        </label>
        <label>
          <input type="radio" name="output_format" value="docx">
          DOCX (Individual)
        </label>
        <label>
          <input type="radio" name="output_format" value="both">
          Both
        </label>
      </div>
      <div class="legend">PDF: All certificates merged into one file · DOCX: Individual files in ZIP</div>
    </div>

    <div class="drop-zone" id="drop-zone">
      <input type="file" name="file" id="file-input" accept=".xlsx,.xls">
      <span class="icon">📊</span>
      <p>Drop your Excel file here</p>
      <p class="hint">or click to browse · .xlsx / .xls</p>
    </div>
    <div id="file-name"></div>

    <div class="info-box">
      <strong>📌 How it works</strong><br>
      Pick the group/sheet to process, then check which certificate type(s) to generate.
      <br><br>
      <strong>Sector = School</strong> rows only ever get the Teachers certificate. Every other row gets
      whichever of VET / Job Shadowing / Portuguese you checked.
      <br><br>
      <strong>Output:</strong> PDF merges all certificates into one file. DOCX gives individual files.
    </div>

    <button type="submit" id="submit-btn" disabled>
      <span id="btn-text">Generate Certificates</span>
    </button>
  </form>

  <div id="progress">
    <div class="bar-wrap"><div class="bar"></div></div>
    <div id="status">Processing…</div>
  </div>

  <div id="result">
    <span class="success-icon">🎉</span>
    <div class="success">Certificates generated successfully!</div>
    <a id="dl-link" class="dl-btn" href="#">
      <span class="icon">⬇</span> Download
    </a>
  </div>

  <div id="error-box"></div>
  </div>

  <footer>
    <p><strong>Wake Up Projects</strong> · Developed by Ariful</p>
  </footer>
</div>

<script>
const dropZone     = document.getElementById('drop-zone');
const fileInput    = document.getElementById('file-input');
const fileName     = document.getElementById('file-name');
const sheetPicker  = document.getElementById('sheet-picker');
const sheetSelect  = document.getElementById('sheet-select');
const submitBtn    = document.getElementById('submit-btn');
const form         = document.getElementById('upload-form');
const progress     = document.getElementById('progress');
const statusEl     = document.getElementById('status');
const resultEl     = document.getElementById('result');
const errorBox     = document.getElementById('error-box');
const dlLink       = document.getElementById('dl-link');

// Portuguese checkbox and hours field
const portugueseCheckbox = document.getElementById('portuguese-cert');
const hoursBox = document.getElementById('portuguese-hours-box');

// Show/hide hours field based on Portuguese checkbox
portugueseCheckbox.addEventListener('change', function() {
    hoursBox.style.display = this.checked ? 'block' : 'none';
});

// Initially show if checked (it is checked by default)
if (portugueseCheckbox.checked) {
    hoursBox.style.display = 'block';
}

// Format radio button styling
document.querySelectorAll('.format-options input[type="radio"]').forEach(rb => {
  rb.addEventListener('change', function() {
    document.querySelectorAll('.format-options label').forEach(l => l.classList.remove('checked'));
    if (this.checked) this.closest('label').classList.add('checked');
  });
  if (rb.checked) rb.closest('label').classList.add('checked');
});

function updateSubmitState() {
  const checked = document.querySelectorAll('input[name="cert_templates"]:checked');
  submitBtn.disabled = !(fileInput.files[0] && sheetSelect.value && checked.length > 0);
}

// Update checkbox styling
document.querySelectorAll('.template-options input[type="checkbox"]').forEach(cb => {
  cb.addEventListener('change', function() {
    this.closest('label').classList.toggle('checked', this.checked);
    updateSubmitState();
  });
  if (cb.checked) cb.closest('label').classList.add('checked');
});

async function loadSheets(file) {
  sheetPicker.style.display = 'block';
  sheetSelect.disabled = true;
  sheetSelect.innerHTML = '<option value="">Loading sheets…</option>';
  submitBtn.disabled = true;
  errorBox.style.display = 'none';

  const fd = new FormData();
  fd.append('file', file);
  try {
    const res = await fetch('/list-sheets', { method: 'POST', body: fd });
    const data = await res.json();
    if (!res.ok) throw new Error(data.error || 'Could not read sheets');
    if (!data.sheets || data.sheets.length === 0) {
      throw new Error('No group sheets found in this file.');
    }
    sheetSelect.innerHTML = '<option value="">— Select a group —</option>' +
      data.sheets.map(s => `<option value="${s.replace(/"/g, '&quot;')}">${s}</option>`).join('');
    sheetSelect.disabled = false;
  } catch (err) {
    sheetSelect.innerHTML = '<option value="">— could not load —</option>';
    errorBox.textContent = '❌ ' + err.message;
    errorBox.style.display = 'block';
  }
  updateSubmitState();
}

sheetSelect.addEventListener('change', updateSubmitState);

fileInput.addEventListener('change', () => {
  if (fileInput.files[0]) {
    fileName.textContent = '📄 ' + fileInput.files[0].name;
    loadSheets(fileInput.files[0]);
  } else {
    fileName.textContent = '';
    updateSubmitState();
  }
});

['dragenter','dragover'].forEach(e => dropZone.addEventListener(e, ev => {
  ev.preventDefault(); dropZone.classList.add('drag-over');
}));
['dragleave','drop'].forEach(e => dropZone.addEventListener(e, ev => {
  ev.preventDefault(); dropZone.classList.remove('drag-over');
}));
dropZone.addEventListener('drop', ev => {
  const f = ev.dataTransfer.files[0];
  if (f) {
    fileInput.files = ev.dataTransfer.files;
    fileName.textContent = '📄 ' + f.name;
    loadSheets(f);
  }
});

form.addEventListener('submit', async e => {
  e.preventDefault();
  errorBox.style.display = 'none';
  resultEl.style.display = 'none';
  progress.style.display = 'block';
  submitBtn.disabled = true;
  document.getElementById('btn-text').textContent = 'Generating…';
  statusEl.textContent = 'Generating certificates…';

  const fd = new FormData(form);
  try {
    const res = await fetch('/generate-certificates', { method: 'POST', body: fd });
    if (!res.ok) {
      const err = await res.json().catch(() => ({ error: 'Unknown error' }));
      throw new Error(err.error || 'Server error');
    }
    const blob = await res.blob();
    const url  = URL.createObjectURL(blob);
    dlLink.href = url;
    dlLink.download = 'certificates.zip';
    progress.style.display = 'none';
    resultEl.style.display = 'block';
    document.getElementById('btn-text').textContent = 'Generate Certificates';
    submitBtn.disabled = false;
  } catch (err) {
    progress.style.display = 'none';
    errorBox.textContent = '❌ ' + err.message;
    errorBox.style.display = 'block';
    document.getElementById('btn-text').textContent = 'Generate Certificates';
    submitBtn.disabled = false;
  }
});

// Initial setup
updateSubmitState();
</script>
</body>
</html>
"""

# Substitute the actual on-disk template filenames into the checkbox values.
HTML = (
    HTML.replace("__VET_FILE__", VET_TEMPLATE_FILE)
        .replace("__JOB_SHADOWING_FILE__", JOB_SHADOWING_FILE)
        .replace("__PORTUGUESE_FILE__", PORTUGUESE_FILE)
        .replace("__TEACHERS_FILE__", TEACHERS_FILE)
)

# Column name candidates, ordered so more specific names are tried first
# where two columns could both match (e.g. "Sector" vs "Sector Database").
COL_MAP = {
    'name': ['name', 'student', 'participant', 'first name', 'nome'],
    'surname': ['surname', 'last name', 'apellido', 'sobrenome'],
    'birth_date': ['date of birth', 'birth date', 'birth_date', 'dob', 'birthday'],
    'start_date': ['start date', 'start_date', 'start', 'begin', 'begin date'],
    'end_date': ['end date', 'end_date', 'end', 'finish', 'finish date'],
    'sector': ['sector database', 'sector', 'area', 'field', 'department', 'study course'],
    'company': ['company', 'host org', 'host_org', 'host organisation', 'organization', 'empresa'],
    'company_address': ['address', 'company address', 'host address', 'host_address', 'endereco', 'direccion'],
    'tutor': ['tutor', 'supervisor', 'mentor', 'contact'],
    'telephone': ['telephone', 'phone', 'contact number', 'tel', 'telefone'],
    'email': ['email', 'e-mail', 'mail', 'correo'],
    'project_code': ['project code', 'project id', 'project number', 'projeto'],
    'notes': ['notes', 'comments', 'observations', 'comentarios', 'feedback'],
    'dress_code': ['dress code', 'uniform', 'attire'],
    'pickup': ['pickup', 'pick up', 'collection'],
    'presentation': ['presentation', 'apresentação', 'start time', 'induction'],
    'age': ['age', 'idade', 'edad'],
    'languages': ['languages', 'idiomas', 'language'],
    'allergies': ['allergies', 'alergies', 'allergens', 'alergias'],
    'experience': ['experience', 'experiencia', 'work experience'],
    'expectation': ['expectation', 'expectations', 'expectativas'],
    'working_hours': ['working hours', 'work hours', 'schedule', 'horario'],
    'vat': ['vat', 'nif', 'tax id', 'vat number'],
    'pic': ['pic', 'participant identification code'],
    'oid': ['oid', 'organisation id'],
    'website': ['website', 'web', 'site', 'pagina web'],
}

# Required header markers used to auto-detect which row is the real header row.
HEADER_MARKERS = ['name', 'surname', 'sector']

# Matches a date range like "17/06/2026 - 02/07/2026" anywhere in a cell.
DATE_RANGE_RE = re.compile(
    r'(?<!\w)(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s*[-–—]\s*(\d{1,2}[./]\d{1,2}[./]\d{2,4})'
)
# Matches an Erasmus+/KA1xx style project code, e.g. "2025-1-ES01-KA121-VET-000339264".
PROJECT_CODE_RE = re.compile(r'\b\d{4}-\d+-[A-Z0-9-]*KA\d[A-Z0-9-]+\b')


def normalize_sector(sector):
    """Clean and normalize sector names for matching."""
    if not sector or pd.isna(sector):
        return ""
    sector = str(sector).strip()
    sector = re.sub(r'\s*[\(\[].*?[\)\]]', '', sector)
    sector = re.sub(r'\s*[-–—]\s*.*$', '', sector)
    return sector.strip().lower()


def find_header_row(sheet_path, sheet_name, max_scan_rows=15, markers=None, min_matches=2):
    """Scan the top of a sheet to find which row actually contains the column headers."""
    if markers is None:
        markers = HEADER_MARKERS
    wb = openpyxl.load_workbook(sheet_path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=max_scan_rows):
            values = [str(c.value).strip().lower() if c.value is not None else '' for c in row]
            matches = sum(1 for marker in markers if any(marker in v for v in values))
            if matches >= min_matches:
                return row[0].row - 1
    finally:
        wb.close()
    return None


def extract_sheet_metadata(sheet_path, sheet_name, header_row_idx, max_scan_rows=15):
    """Pull a shared date range and/or project code from the title rows."""
    metadata = {'start_date': '', 'end_date': '', 'project_code': ''}
    if header_row_idx is None:
        scan_limit = max_scan_rows
    else:
        scan_limit = header_row_idx + 1

    wb = openpyxl.load_workbook(sheet_path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=scan_limit):
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value)
                date_match = DATE_RANGE_RE.search(text)
                if date_match and not metadata['start_date']:
                    metadata['start_date'] = date_match.group(1)
                    metadata['end_date'] = date_match.group(2)
                code_match = PROJECT_CODE_RE.search(text)
                if code_match and not metadata['project_code']:
                    metadata['project_code'] = code_match.group(0)
    finally:
        wb.close()
    return metadata


def match_column(df_columns, possible):
    """Find the best-matching column name for a standard field."""
    best_col, best_score = None, None
    for col in df_columns:
        col_lower = col.lower().strip()
        col_score = None
        for priority, term in enumerate(possible):
            if col_lower == term:
                score = (0, priority)
            elif term in col_lower:
                score = (1, priority)
            else:
                continue
            if col_score is None or score < col_score:
                col_score = score
        if col_score is not None and (best_score is None or col_score < best_score):
            best_score = col_score
            best_col = col
    return best_col


def extract_student_data(df):
    """Extract student/group-row data from a dataframe."""
    students = []

    actual_cols = {}
    for standard, possible in COL_MAP.items():
        col = match_column(df.columns, possible)
        if col is not None:
            actual_cols[standard] = col

    for idx, row in df.iterrows():
        name_col = actual_cols.get('name', 'name')
        surname_col = actual_cols.get('surname', 'surname')

        if name_col not in df.columns and surname_col not in df.columns:
            continue

        first_name = str(row.get(name_col, '')).strip() if name_col in df.columns else ''
        surname = str(row.get(surname_col, '')).strip() if surname_col in df.columns else ''

        if (not first_name or first_name.lower() == 'nan') and (not surname or surname.lower() == 'nan'):
            continue

        student = {}

        if first_name and first_name.lower() != 'nan':
            if surname and surname.lower() != 'nan':
                student['name'] = f"{first_name} {surname}"
            else:
                student['name'] = first_name
        elif surname and surname.lower() != 'nan':
            student['name'] = surname
        else:
            continue

        for standard, col in actual_cols.items():
            if standard not in ['name', 'surname']:
                if col in df.columns:
                    val = row.get(col, '')
                    if pd.isna(val):
                        val = ''
                    elif isinstance(val, (pd.Timestamp, datetime)):
                        val = val.strftime('%d/%m/%Y')
                    student[standard] = str(val).strip()
                else:
                    student[standard] = ''

        for key in ['birth_date', 'start_date', 'end_date', 'sector', 'company', 'company_address',
                    'tutor', 'telephone', 'email', 'project_code', 'notes', 'dress_code', 'pickup',
                    'presentation', 'age', 'languages', 'allergies', 'experience', 'expectation',
                    'working_hours', 'vat', 'pic', 'oid', 'website']:
            student.setdefault(key, '')

        student.setdefault('sector_database', student.get('sector', ''))

        students.append(student)

    return students


# Sheet names that are never cohorts of students.
NON_COHORT_SHEET_NAMES = {
    'task database', 'company database', 'task database - olivetrain',
}


def is_cohort_sheet(sheet_name):
    """True if a sheet looks like a student/group sheet rather than an admin sheet."""
    name_lower = sheet_name.strip().lower()
    if name_lower in NON_COHORT_SHEET_NAMES:
        return False
    if name_lower.startswith('template'):
        return False
    return True


def list_templates():
    """Scan the templates/ folder for .docx files."""
    try:
        files = sorted(f for f in os.listdir(TEMPLATES_FOLDER) if f.lower().endswith('.docx'))
    except FileNotFoundError:
        files = []
    templates = []
    for f in files:
        label = os.path.splitext(f)[0]
        label = re.sub(r'[_-]+', ' ', label).strip().title()
        templates.append({"file": f, "label": label})
    return templates


@app.route("/")
def home():
    return render_template_string(HTML)


@app.route("/list-sheets", methods=["POST"])
def list_sheets():
    """Return the cohort/group sheet names in an uploaded workbook."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded."}), 400

    excel_file = request.files["file"]
    if excel_file.filename == "":
        return jsonify({"error": "No file selected."}), 400

    excel_path = save_upload(excel_file)

    try:
        xl = pd.ExcelFile(excel_path)
        sheets = [s for s in xl.sheet_names if is_cohort_sheet(s)]
    except Exception as e:
        return jsonify({"error": f"Could not read Excel file: {e}"}), 400
    finally:
        try:
            os.remove(excel_path)
        except Exception:
            pass

    return jsonify({"sheets": sheets})


@app.route("/generate-certificates", methods=["POST"])
def generate_certificates():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded."}), 400

    excel_file = request.files["file"]
    if excel_file.filename == "":
        return jsonify({"error": "No file selected."}), 400

    sheet_name_filter = request.form.get('sheet_name', '').strip()
    if not sheet_name_filter:
        return jsonify({"error": "Please choose a group/sheet to generate certificates for."}), 400

    selected_templates = request.form.getlist("cert_templates")
    if not selected_templates:
        return jsonify({"error": "Please select at least one certificate type."}), 400

    # Get output format preference
    output_format = request.form.get('output_format', 'pdf')
    if output_format not in ['pdf', 'docx', 'both']:
        output_format = 'pdf'

    # Get Portuguese hours (default to 6 if not provided or invalid)
    try:
        portuguese_hours = request.form.get('portuguese_hours', '6')
        portuguese_hours = int(portuguese_hours)
        if portuguese_hours < 1:
            portuguese_hours = 6
    except (ValueError, TypeError):
        portuguese_hours = 6

    # Validate every selected template actually exists on disk.
    available_templates = {t['file'] for t in list_templates()}
    for t in selected_templates:
        if t not in available_templates:
            return jsonify({"error": f"Template '{t}' not found in templates folder."}), 400

    excel_path = save_upload(excel_file)

    try:
        xl = pd.ExcelFile(excel_path)
        all_students = []

        for sheet_name in xl.sheet_names:
            if sheet_name.lower() != sheet_name_filter.lower():
                continue
            if not is_cohort_sheet(sheet_name):
                continue

            try:
                header_row_idx = find_header_row(excel_path, sheet_name)
                if header_row_idx is None:
                    continue

                df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header_row_idx)
                df.columns = [str(col).strip() for col in df.columns]

                if len(df) == 0:
                    continue

                has_name = any('name' in str(col).lower() or 'surname' in str(col).lower() for col in df.columns)
                has_sector = any('sector' in str(col).lower() or 'study' in str(col).lower() for col in df.columns)

                if not (has_name or has_sector):
                    continue

                sheet_meta = extract_sheet_metadata(excel_path, sheet_name, header_row_idx)

                students = extract_student_data(df)
                for s in students:
                    s['sheet_name'] = sheet_name
                    if not s.get('start_date') and sheet_meta['start_date']:
                        s['start_date'] = sheet_meta['start_date']
                    if not s.get('end_date') and sheet_meta['end_date']:
                        s['end_date'] = sheet_meta['end_date']
                    if not s.get('project_code') and sheet_meta['project_code']:
                        s['project_code'] = sheet_meta['project_code']
                all_students.extend(students)
            except Exception as e:
                print(f"Error processing sheet {sheet_name}: {e}")
                continue

        if not all_students:
            return jsonify({"error": f"No student data found in sheet '{sheet_name_filter}'. Make sure it has columns like 'Name' or 'Surname'."}), 400

        # Build a sector -> {detailed_tasks, job_related_competences} map from the Task Database sheet.
        sector_map = {}
        try:
            task_sheet_names = ['Task Database', 'TASK DATABASE', 'Task database', 'tasks', 'Sectors']
            task_df = None
            for name in task_sheet_names:
                if name not in xl.sheet_names:
                    continue
                try:
                    task_header_idx = find_header_row(
                        excel_path, name,
                        markers=['sector', 'detailed', 'task', 'competence', 'knowledge'],
                        min_matches=2,
                    )
                    if task_header_idx is None:
                        task_header_idx = 0
                    task_df = pd.read_excel(excel_path, sheet_name=name, header=task_header_idx)
                    break
                except Exception:
                    continue

            if task_df is not None:
                task_df.columns = [str(col).strip() for col in task_df.columns]

                sector_col = None
                tasks_col = None
                competences_col = None

                for col in task_df.columns:
                    col_lower = col.lower()
                    if 'sector' in col_lower:
                        sector_col = col
                    elif 'detailed' in col_lower or 'task' in col_lower or 'Detailed tasks' in col:
                        tasks_col = col
                    elif 'job' in col_lower or 'competence' in col_lower or 'knowledge' in col_lower or 'Job-related' in col:
                        competences_col = col

                if sector_col and tasks_col:
                    for _, row in task_df.iterrows():
                        sector = str(row.get(sector_col, "")).strip()
                        if sector and sector.lower() != "nan":
                            sector_key = normalize_sector(sector)
                            sector_map[sector_key] = {
                                "detailed_tasks": str(row.get(tasks_col, "") or ""),
                                "job_related_competences": str(row.get(competences_col, "") or "") if competences_col else "",
                            }
        except Exception:
            pass

    except Exception as e:
        return jsonify({"error": f"Could not read Excel file: {e}"}), 400

    try:
        os.remove(excel_path)
    except Exception:
        pass

    generated_files = []
    pdf_files = []
    errors = []

    # Create a temporary directory for PDF generation
    temp_dir = tempfile.mkdtemp()

    try:
        for student in all_students:
            try:
                notes = student.get('notes', '').lower()
                experience = student.get('experience', '').lower()
                if 'will not come' in notes or "didn't show up" in notes or 'will not come' in experience or "didn't show up" in experience:
                    continue

                sector_key = normalize_sector(student.get('sector', ''))
                is_school = sector_key == 'school'

                # School-sector rows ONLY ever get the Teachers certificate.
                if is_school:
                    templates_for_row = [t for t in selected_templates if t == TEACHERS_FILE]
                else:
                    templates_for_row = [t for t in selected_templates if t != TEACHERS_FILE]

                if not templates_for_row:
                    continue

                ctx = {
                    "NAME": student.get('name', ''),
                    "BIRTH_DATE": student.get('birth_date', ''),
                    "START": student.get('start_date', ''),
                    "END": student.get('end_date', ''),
                    "SECTOR": student.get('sector', ''),
                    "SECTOR_DATABASE": student.get('sector_database', ''),
                    "HOST_ORG": student.get('company', ''),
                    "HOST_ORG_ADDRESS": student.get('company_address', ''),
                    "PROJECT_CODE": student.get('project_code', ''),
                    "TUTOR": student.get('tutor', ''),
                    "TELEPHONE": student.get('telephone', ''),
                    "EMAIL": student.get('email', ''),
                    "NOTES": student.get('notes', ''),
                    "DRESS_CODE": student.get('dress_code', ''),
                    "PICKUP": student.get('pickup', ''),
                    "PRESENTATION": student.get('presentation', ''),
                    "WORKING_HOURS": student.get('working_hours', ''),
                    "AGE": student.get('age', ''),
                    "LANGUAGES": student.get('languages', ''),
                    "ALLERGIES": student.get('allergies', ''),
                    "EXPERIENCE": student.get('experience', ''),
                    "EXPECTATION": student.get('expectation', ''),
                    "VAT": student.get('vat', ''),
                    "PIC": student.get('pic', ''),
                    "OID": student.get('oid', ''),
                    "WEBSITE": student.get('website', ''),
                    "SHEET_NAME": student.get('sheet_name', ''),
                    "PORTUGUESE_HOURS": str(portuguese_hours),
                    "detailed_tasks": "",
                    "job_related_competences": "",
                }

                if sector_key in sector_map:
                    ctx["detailed_tasks"] = sector_map[sector_key].get("detailed_tasks", "")
                    ctx["job_related_competences"] = sector_map[sector_key].get("job_related_competences", "")
                else:
                    for key in sector_map:
                        if key in sector_key or sector_key in key:
                            ctx["detailed_tasks"] = sector_map[key].get("detailed_tasks", "")
                            ctx["job_related_competences"] = sector_map[key].get("job_related_competences", "")
                            break

                safe_name = re.sub(r'[^\w\s-]', '', student.get('name', 'student'))
                safe_name = re.sub(r'[-\s]+', '_', safe_name)

                for template_file in templates_for_row:
                    template_path = os.path.join(TEMPLATES_FOLDER, template_file)
                    doc = DocxTemplate(template_path)
                    doc.render(ctx)
                    suffix = TEMPLATE_SUFFIX.get(template_file, os.path.splitext(template_file)[0])
                    out_path = os.path.join(OUTPUT_FOLDER, f"{safe_name}_{suffix}.docx")
                    doc.save(out_path)
                    generated_files.append(out_path)

            except Exception as e:
                errors.append(f"Error processing {student.get('name', 'Unknown')}: {str(e)}")

        if not generated_files:
            error_msg = "No certificates generated."
            if errors:
                error_msg += " Errors: " + "; ".join(errors[:3])
            error_msg += f" No matching rows found in '{sheet_name_filter}' for the selected certificate type(s)."
            return jsonify({"error": error_msg}), 400

        # Convert every generated DOCX to PDF in ONE LibreOffice batch call —
        # much faster than converting one file at a time, and it works the
        # same way on Windows, Linux, and Mac.
        if output_format in ['pdf', 'both']:
            try:
                docx_to_pdf = convert_docx_to_pdf_batch(generated_files, temp_dir)
                pdf_files = [docx_to_pdf[f] for f in generated_files if f in docx_to_pdf]
                missing = len(generated_files) - len(pdf_files)
                if missing:
                    errors.append(f"{missing} file(s) failed to convert to PDF.")
            except Exception as e:
                errors.append(f"PDF conversion failed: {e}")
                if output_format == 'pdf':
                    return jsonify({"error": f"PDF conversion failed: {e}"}), 500

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            if output_format in ['pdf', 'both'] and pdf_files:
                # Merge all PDFs into one
                merged_pdf_path = os.path.join(temp_dir, "merged_certificates.pdf")
                if merge_pdfs(pdf_files, merged_pdf_path):
                    zf.write(merged_pdf_path, "certificates_merged.pdf")
                else:
                    # If merge fails, add individual PDFs
                    for pdf in pdf_files:
                        if os.path.exists(pdf):
                            zf.write(pdf, os.path.basename(pdf))

            if output_format in ['docx', 'both']:
                # Add individual DOCX files
                for f in generated_files:
                    if os.path.exists(f):
                        zf.write(f, os.path.basename(f))

        zip_buffer.seek(0)

        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name="certificates.zip",
            mimetype="application/zip"
        )

    finally:
        # Clean up generated files
        for f in generated_files:
            try:
                os.remove(f)
            except Exception:
                pass
        # Clean up temp directory
        try:
            shutil.rmtree(temp_dir)
        except Exception:
            pass


if __name__ == "__main__":
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(debug=debug_mode, port=5000)